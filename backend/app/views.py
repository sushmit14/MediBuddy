from django.shortcuts import render
from adrf.views import APIView
from .models import *
from rest_framework.response import Response
from .serializer import *
from django.http import JsonResponse
from django.http import HttpResponse
from datetime import datetime
import json
import random
from PyPDF2 import PdfReader
from docx import Document
from pptx import Presentation
from openpyxl import load_workbook
from langchain.document_loaders.text import TextLoader
from langchain.indexes import VectorstoreIndexCreator
from langchain.indexes.vectorstore import VectorStoreIndexWrapper
from langchain.vectorstores.chroma import Chroma
from langchain.chat_models import ChatOpenAI
from langchain.embeddings.openai import OpenAIEmbeddings
import os
import string
from tqdm import tqdm 
import pickle
from openai import OpenAI
from nltk.corpus import wordnet as wn
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import cloudinary
import cloudinary.uploader
from django.http import JsonResponse
from rest_framework.views import APIView
import os


os.environ['OPENAI_API_KEY'] = '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'
client = OpenAI(api_key=os.environ['OPENAI_API_KEY'])
storeContext = True
class ReactView(APIView):

    serializer_class = ReactSerializer

    def get(self, request):
        random_number = random.randint(1, 100)
        print(request.body)
        # Print a message with timestamp in the backend
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"Received GET request at {timestamp}. Random number: {random_number}")

        # Respond with a JSON containing the random number
        return JsonResponse({"random_number": random_number})

    def post(self, request):

        serializer = ReactSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response(serializer.data)

class FindSum(APIView):
    serializer_class = ReactSerializer

    def post(self, request):
        # random_number = random.randint(1, 100)
        try:
            # Deserialize the JSON data from the request body
            data = json.loads(request.body.decode('utf-8'))
            
            # Access the values 'a' and 'b' from the JSON data
            a = data.get('a', 0)
            b = data.get('b', 0)

            # Perform the operation
            c = a + b

            # Print a message with timestamp in the backend
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"Received post request at {timestamp}. Result of sum: {c}")

            # Respond with a JSON containing the result
            return JsonResponse({"sum_result": c})

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format in the request body"}, status=400)

    # def post(self, request):

    #     serializer = ReactSerializer(data=request.data)
    #     if serializer.is_valid(raise_exception=True):
    #         serializer.save()
    #         return Response(serializer.data)


class Upload(APIView):
    def post(self, request):
        try:
            
            cloudinary.config( 
            cloud_name = "dqln5phax", 
            api_key = "373247933775311", 
            api_secret = "waEbZ9xlEpjAp8PhMfP6ESw3ZCs" 
            )
            # os.environ['CLOUDINARY_URL'] = "cloudinary://373247933775311:waEbZ9xlEpjAp8PhMfP6ESw3ZCs@dqln5phax"
            if 'file' not in request.FILES:
                return JsonResponse({"error": "No file found in request."}, status=400)
            
            file_to_upload = request.FILES['file']
            
            # Upload the file to Cloudinary
            uploaded_file = cloudinary.uploader.upload(file_to_upload, use_filename=True,access_mode='public')
            storeContext = False
            # Get the URL of the uploaded file from Cloudinary response
            uploaded_file_url = uploaded_file['secure_url']
            
            # You can do further processing or saving of this URL if needed
            
            return JsonResponse({"message": "PDF file uploaded to Cloudinary", "file_url": uploaded_file})
        except Exception as e:
            print(e)
            return JsonResponse({"error": str(e)}, status=400)
        
class UploadPdf(APIView):
    # index = None  # Class attribute to store the index

    def post(self, request):
        try:
            # Get the uploaded file from the request
            uploaded_file = request.FILES['file']

            # Function to filter out unwanted characters
            # def clean_text(raw_text):
            #     printable = set(string.printable)
            #     return ''.join(filter(lambda x: x in printable, raw_text))

            # progress = 0  # Initialize progress variable

            # # Extract text based on file type
            # if uploaded_file.name.endswith('.pdf'):
            #     # For PDF files
            #     reader = PdfReader(uploaded_file)
            #     text = ""
            #     total_pages = len(reader.pages)
            #     for i in tqdm(range(total_pages), desc="Extracting PDF Text"):
            #         text += clean_text(reader.pages[i].extract_text())
            #         # Update progress
            #         progress = int((i + 1) / total_pages * 100)
            # elif uploaded_file.name.endswith('.xlsx'):
            #     # For Excel files
            #     wb = load_workbook(uploaded_file)
            #     text = ""
            #     for sheet in wb.sheetnames:
            #         ws = wb[sheet]
            #         for row in ws.iter_rows():
            #             for cell in row:
            #                 if cell.value:
            #                     text += clean_text(str(cell.value)) + " "
            # elif uploaded_file.name.endswith('.txt'):
            #     # For Text files
            #     text = ""
            #     with open(uploaded_file, 'r') as file:
            #         text += file.read()
            # Save the extracted text to a text file
            save_path = "D:\Sushmit Dasgupta\Web Development\django-proj\project\Drug_Disease.txt"  # Change this to the desired path
            # with open(save_path, 'w', encoding='utf-8') as text_file:
            #     text_file.write(text)

            loader = TextLoader(save_path)
            print("creating index")
            index = VectorstoreIndexCreator(vectorstore_kwargs={ "persist_directory": "D:/Sushmit Dasgupta/Web Development/django-proj/project/app/disease_40k"}).from_loaders([loader])
            print("index created")
            # Create index after uploading the file if it hasn't been created yet
            # if not UploadPdf.index:
            #     loader = TextLoader(save_path)
            #     UploadPdf.index = VectorstoreIndexCreator().from_loaders([loader])

            response = JsonResponse({"message": "File uploaded successfully."})
            response["Access-Control-Allow-Origin"] = "http://localhost:3005"
            response["Access-Control-Allow-Credentials"] = "true"
            
            return JsonResponse({"message": "PDF file uploaded, text extracted, and index created successfully"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

queries = []     
class SearchQuery(APIView):
    # def __init__(self):
        # self.queries = []
        # self.q=0
    def clean_text(self,raw_text):
            printable = set(string.printable)
            return ''.join(filter(lambda x: x in printable, raw_text))
    def fun1(self):
        openai_embeddings = OpenAIEmbeddings()
        # if is_disease_to_drug_mapping:           
        persistent_index_path = "D:/Sushmit Dasgupta/Web Development/django-proj/project/app/disease_40k"
        # else:
        #     persistent_index_path = "D:/Sushmit Dasgupta/Web Development/django-proj/project/app/symptoms"
        index = VectorStoreIndexWrapper(vectorstore=Chroma(embedding_function = openai_embeddings, persist_directory=persistent_index_path))
        return index
    
    def fun2(self):
        openai_embeddings = OpenAIEmbeddings()
        persistent_index_path = "D:/Sushmit Dasgupta/Web Development/django-proj/project/app/symptoms"
        index = VectorStoreIndexWrapper(vectorstore=Chroma(embedding_function = openai_embeddings, persist_directory=persistent_index_path))
        return index
    
    def is_string_contained(self,main_string, string_array):
        for string in string_array:
            if string in main_string:
                return True
        return False
    
    def jaccard_similarity(self,word, string):
        word_set = set(word)
        string_set = set(string.split())
        
        intersection = len(word_set.intersection(string_set))
        union = len(word_set.union(string_set))
        
        similarity = intersection / union if union != 0 else 0.0
        return similarity
    
    def preprocess_text(self,text):
    # Tokenize the text
        tokens = word_tokenize(text)
        
        # Remove stopwords
        stop_words = set(stopwords.words('english'))
        filtered_tokens = [token for token in tokens if token.lower() not in stop_words]
        
        # Lemmatize tokens
       
        lemmatizer = WordNetLemmatizer()
        lemmatized_tokens = [lemmatizer.lemmatize(token.lower()) for token in filtered_tokens]
        
        return lemmatized_tokens

    def wordnet_similarity(self,word, string):
        word_synsets = wn.synsets(word)
        string_synsets = [synset for token in self.preprocess_text(string) for synset in wn.synsets(token)]
        
        max_similarity = 0.0
        
        for word_synset in word_synsets:
            for string_synset in string_synsets:
                similarity = word_synset.path_similarity(string_synset)
                if similarity is not None and similarity > max_similarity:
                    max_similarity = similarity
        
        return max_similarity
    
    def find_limits(self,json_text):
        """
        Find upper and lower limits for each test result in the given JSON text.

        Args:
        json_text (str): JSON text containing test results.

        Returns:
        dict: A dictionary containing test names as keys and their corresponding upper and lower limits as values.
        """
        limits = {}

        # Parse the JSON text
        data = json.loads(json_text)

        # Iterate over TestResult items
        for result in data["TestResult"]:
            name = result.get("name")
            upper_limit = result.get("normal_range_upper_limit")
            lower_limit = result.get("normal_range_lower_limit")

            # Check if both upper and lower limits are available
            if upper_limit is not None and lower_limit is not None:
                limits[name] = (lower_limit, upper_limit)

        return limits
    
    def post(self, request):
        try:
            # Clear the queries list and reset q to 0
            queries = []
            print("Query list emptied")
            return HttpResponse("Queries list cleared and q reset to 0")
        except Exception as e:
            return HttpResponse(f"Error clearing queries list and resetting q: {str(e)}")
        
    def get(self, request):
        openai_embeddings = OpenAIEmbeddings()
        
        try:
            query = request.GET.get('query', '')

            # custom_functions = [
            #     {
            #         'name': 'fun1',
            #         'description': 'only if asks for medicines',
            #         'parameters': {
            #         }
            #     },
            #     {
            #         'name': 'fun2',
            #         'description': 'only if contains symptoms',
            #         'parameters': {}
            #     }
            # ]
            word = "symptom"
            similarity = self.jaccard_similarity(word,query)
            print(similarity)
            querypr = query + " If the above query is not a medical question, strictly return only 1."
            medicine_syn = ['medicine','drug','medication','treatment','cure','remedy', 'medicament']
            if self.is_string_contained(query.lower(), medicine_syn): 
                index = self.fun1()
                query += " Provide DrugBank IDs and answer confidently. "
                print("diseases")
            else:
                index = self.fun2()
                query += " Find mild diseases having the given subset of symptoms. If no mild diseases match, only then check other diseases."
                print("symptoms")

           
            print(querypr)
            response = client.chat.completions.create(
                model = 'gpt-3.5-turbo',
                messages = [{'role': 'user', 'content': querypr}],
                # functions = custom_functions,
                # function_call = 'auto'
            )

            print(response.choices[0])
            # available_functions = {
            #     "fun1": self.fun1,
            #     "fun2": self.fun2
            # }
            # function_to_call = available_functions[response.choices[0].message.function_call.name]
            if response.choices[0].message.content == '0':
                response = "This query is out of scope for the current chatbot."
                return JsonResponse({"response": response})
            #     index = self.fun2()
            #     print("Index for symptoms created")
            # elif response.choices[0].message.content == '0':
            #     index = self.fun1()
            # else:
            #     response = "This query is out of scope for the current chatbot."
            #     return JsonResponse({"response": response})

            json_text = """
            {
            "MedInfo": [
                {
                    "name": "Paracetamol",
                    "frequency": 3,
                    "number_of_days": 5,
                    "specific_part_of_day": "after meals"
                }
            ],
            "PatientInfo": [],
            "TestResult": [
                {
                    "name": "Test Name",
                    "value": null,
                    "unit": "Units",
                    "normal_range_lower_limit": null,
                    "normal_range_upper_limit": null,
                    "flag": ""
                },
                {
                    "name": "AST (SGOT)",
                    "value": 35.0,
                    "unit": "U/L",
                    "normal_range_lower_limit": null,
                    "normal_range_upper_limit": 40,
                    "flag": null
                },
                {
                    "name": "ALT (SGPT)",
                    "value": 56.0,
                    "unit": "U/L",
                    "normal_range_lower_limit": 0,
                    "normal_range_upper_limit": 41,
                    "flag": "Abnormal"
                },
                {
                    "name": "IFCC without P5P"
                },
                {
                    "name": "AST:ALT Ratio",
                    "value": 0.63,
                    "unit": "",
                    "normal_range_lower_limit": 0,
                    "normal_range_upper_limit": 1,
                    "flag": "<1.00"
                },
                {
                    "name": "GGTP",
                    "value": 59.0,
                    "unit": "U/L",
                    "normal_range_lower_limit": 0,
                    "normal_range_upper_limit": 71.0,
                    "flag": "Normal"
                },
                {
                    "name": "Complete Blood Count",
                    "value": 1,
                    "unit": "CBC",
                    "normal_range_lower_limit": 0,
                    "normal_range_upper_limit": 1,
                    "flag": "Normal"
                },
                {
                    "name": "Hemoglobin",
                    "value": 16.0,
                    "unit": "g/dL",
                    "normal_range_lower_limit": 13.0,
                    "normal_range_upper_limit": 17.0,
                    "flag": "Normal"
                },
                {
                    "name": "Packed Cell Volume (PCV)",
                    "value": 48.0,
                    "unit": "%",
                    "normal_range_lower_limit": 40.0,
                    "normal_range_upper_limit": 50.0,
                    "flag": "Normal"
                },
                {
                    "name": "RBC Count",
                    "value": 5.08,
                    "unit": "mill/mm3",
                    "normal_range_lower_limit": 4.5,
                    "normal_range_upper_limit": 5.5
                },
                {
                    "name": "MCV",
                    "value": 94.5,
                    "unit": "fL",
                    "normal_range_lower_limit": 83.0,
                    "normal_range_upper_limit": 101.0
                },
                {
                    "name": "MCH",
                    "value": 31.4,
                    "unit": "pg",
                    "normal_range_lower_limit": 27.0,
                    "normal_range_upper_limit": 32.0
                },
                {
                    "name": "MCHC",
                    "value": 33.2,
                    "unit": "g/dL",
                    "normal_range_lower_limit": 31.5,
                    "normal_range_upper_limit": 34.5,
                    "flag": "Normal"
                },
                {
                    "name": "Red Cell Distribution Width (RDW)",
                    "value": 14.2,
                    "unit": "%",
                    "normal_range_lower_limit": 11.6,
                    "normal_range_upper_limit": 14.0,
                    "flag": "Abnormal"
                },
                {
                    "name": "Total Leukocyte Count",
                    "value": 6.6,
                    "unit": "thou/mm3",
                    "normal_range_lower_limit": 4.0,
                    "normal_range_upper_limit": 10.0
                },
                {
                    "name": "Segmented Neutrophils",
                    "value": 57.9,
                    "unit": "%",
                    "normal_range_lower_limit": 40,
                    "normal_range_upper_limit": 80,
                    "flag": "Normal"
                },
                {
                    "name": "Lymphocytes",
                    "value": 25.0,
                    "unit": "%",
                    "normal_range_lower_limit": 20.0,
                    "normal_range_upper_limit": 40.0,
                    "flag": "Normal"
                },
                {
                    "name": "Monocytes",
                    "value": 10.3,
                    "unit": "%",
                    "normal_range_lower_limit": 2.0,
                    "normal_range_upper_limit": 10.0,
                    "flag": "Normal"
                },
                {
                    "name": "Eosinophils",
                    "value": 6.7,
                    "unit": "%",
                    "normal_range_lower_limit": 1.0,
                    "normal_range_upper_limit": 6.0,
                    "flag": "Abnormal"
                },
                {
                    "name": "Basophils",
                    "value": 0.1,
                    "unit": "%",
                    "normal_range_lower_limit": 0,
                    "normal_range_upper_limit": 2,
                    "flag": "Abnormal"
                },
                {
                    "name": "Absolute Leucocyte Count"
                },
                {
                    "name": "Neutrophils",
                    "value": 3.82,
                    "unit": "thou/mm3",
                    "normal_range_lower_limit": 2.0,
                    "normal_range_upper_limit": 7.0,
                    "flag": "Normal"
                },
                {
                    "name": "Lymphocytes",
                    "value": 1.65,
                    "unit": "thou/mm3",
                    "normal_range_lower_limit": 1.0,
                    "normal_range_upper_limit": 3.0
                },
                {
                    "name": "Monocytes",
                    "value": 0.68,
                    "unit": "thou/mm3",
                    "normal_range_lower_limit": 0.2,
                    "normal_range_upper_limit": 1.0,
                    "flag": "Normal"
                },
                {
                    "name": "Eosinophils",
                    "value": 0.44,
                    "unit": "thou/mm3",
                    "normal_range_lower_limit": 0.02,
                    "normal_range_upper_limit": 0.5
                },
                {
                    "name": "Basophils",
                    "value": 0.01,
                    "unit": "thou/mm3",
                    "normal_range_lower_limit": 0.02,
                    "normal_range_upper_limit": 0.1
                },
                {
                    "name": "Platelet Count",
                    "value": 222.0,
                    "unit": "thou/mm3",
                    "normal_range_lower_limit": 150.0,
                    "normal_range_upper_limit": 410.0,
                    "flag": "Normal"
                },
                {
                    "name": "Mean Platelet Volume",
                    "value": 9.8,
                    "unit": "fL",
                    "normal_range_lower_limit": 6.5,
                    "normal_range_upper_limit": 12.0
                }
            ]
        }
            """

            # test_limits = self.find_limits(query)
            # dict_string = json.dumps(test_limits)
            # query2 = dict_string + " Identify which disease."

            # print(test_limits.join())
            query2 = ""
            cnt = 0
            for i in queries:
                # print(i, type(i))
                if cnt %2 == 0:
                    query2 = query2 + "Question: "+i
                else:
                    query2 = query2 + "Answer: "+i
                cnt += 1
            # if (len(self.queries) == 0):
            #     query2 = "Do not use any context from previous chats. Treat this as fresh."
            query2 += query
            # print(query2)
            print(f"Received query: '{query}'")
            save_path = "D:\Sushmit Dasgupta\Web Development\django-proj\project\extracted_text.txt"  # Change this to the desired path
            # save_path_index = "D:\Sushmit Dasgupta\Web Development\django-proj\project" 
            # loader = TextLoader(save_path)
            # index = VectorstoreIndexCreator().from_loaders([loader])
            # persistent_index_path = "D:/Sushmit Dasgupta/Web Development/django-proj/project/app/disease_40k"
            # index = VectorStoreIndexWrapper(vectorstore=Chroma(embedding_function = openai_embeddings, persist_directory=persistent_index_path))

            # symptoms_path =  "D:/Sushmit Dasgupta/Web Development/django-proj/project/app/symptoms"
            # index_symptoms = VectorStoreIndexWrapper(vectorstore=Chroma(embedding_function = openai_embeddings, persist_directory=symptoms_path))
            response2 = index.query(query2,llm=ChatOpenAI(model="gpt-4-1106-preview"))
            # print(response)
            if len(queries) == 3:
                queries.pop(0)
                queries.pop(0)
            # print(type(response))
            print("\n\n\n", storeContext, "\n\n\n")
            if(storeContext):
                queries.append(query)
                queries.append(str(self.clean_text(response2)))
            print(queries)
            return JsonResponse({"response": response2})
            # else:
            #     print("Index not available. Upload a file first.")
            #     return JsonResponse({"error": "Index not available. Upload a file first."}, status=400)

        except Exception as e:
            print("Exception:", e)
            return JsonResponse({"error": str(e)}, status=400)
        
def testing(self, request):
    print("Testing the Django backend")
    print(request.data)
    return Response({"message": "Testing successful"})